# Thomas Berger
# 10/10/2020

'''
Create an Excel with colored cells recreating an image.

Put in a folder all photos (png or jpg) you want to transform and this file then run it
It will output a excel file for each photo.


''' 

# Risizing image to this maximum sieze
# Prevent excel for having a style formatting error when to much cells are painted
max_size = 300

import os
from openpyxl.styles import  PatternFill
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from PIL import Image

#Get all image files in directory
path = os.getcwd()
files = os.listdir(path)
matchers = ['.jpg','.png']
matching = [s for s in files if any(xs in s for xs in matchers)]

#for each image file 
for k in matching:
    # Get image and resize
    im = Image.open(k)
    maxwidth = max_size
    maxheight = max_size 
    
    if im.size>=(maxwidth,maxheight):
        ratio = min(maxwidth/im.size[0], maxheight/im.size[1])
        im = im.resize((round(im.size[0]*ratio),round(im.size[1]*ratio)))
    pix = im.load()
    
    # Create workbook, worksheet
    wb = Workbook()
    ws = wb.active
    
    #scann all pixels
    for i in range(1,im.size[0]):
        for j in range(1,im.size[1]):
            #extract color and put it in Hex
            color = pix[i,j]
            t1 = "0x{:02x}".format(color[0])
            t2 = "0x{:02x}".format(color[1])
            t3 = "0x{:02x}".format(color[2])
            
            # Paint cell
            ws.cell(row=j, column=i).fill = PatternFill("solid", fgColor=t1[2:4]+t2[2:4]+t3[2:4])
            
            # Set cell squares
            ws.row_dimensions[i].height = 10
            ws.column_dimensions[get_column_letter(j)].width = 2
        
        # Print advancement
        perc = i/im.size[0]*100
        print('%.2f%%' % perc , flush=True)
    
    
    #Save workbook
    print('Saving workbook ...')
    wb.save("Painted%s.xlsx" % k[:-4])
    print('Finished')